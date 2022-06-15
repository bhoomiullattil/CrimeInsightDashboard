from bs4 import BeautifulSoup
import urllib.request,sys,time
import requests 
import pandas as pd
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import locationtagger
import xlrd
import openpyxl
locations_City=[]
Statement=[]
state_name=[]
pg=0
sub=str(pg)
url="https://www.indiatoday.in/crime?page="+sub
pg+=1
try:
	page=requests.get(url)
except:
	print("Something wrong")
soup = BeautifulSoup(page.text, "html.parser")
links=soup.find_all('div',attrs={'class':'catagory-listing'})
Statement_vals = soup.find_all("div",attrs={'class':'detail'})
for S_val in Statement_vals:
	Statement.append(S_val.text.strip())

for sentence in Statement:
	place_entity = locationtagger.find_locations(text = sentence)
	locations_City.extend(place_entity.cities)
	
loc=("state_city.xls")
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
col=0
state_col=1
for city in locations_City:
	for row in range(sheet.nrows):
		if sheet.cell(row,col).value==city:
			state_name.append([city,sheet.cell(row,state_col).value])
			break
#print(state_name)
wb = openpyxl.load_workbook("stored_data.xlsx")
sheet = wb.active
rows=sheet.max_row

for elements in state_name:
	column=1
	rows=rows+1
	c1=sheet.cell(row=rows,column=column)
	c1.value=elements[0]
	c2=sheet.cell(row=rows,column=column+1)
	c2.value=elements[1]
	
	wb.save("stored_data.xlsx")
print(rows)









